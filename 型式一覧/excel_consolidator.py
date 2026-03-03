"""
Excelファイル統合スクリプト
指定フォルダ内のExcelファイルから特定のデータを抽出し、CSVに出力します。
"""

import csv
import warnings
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
import xlrd

# =====================================
# 設定セクション - ここで項目を追加/変更できます
# =====================================

# モード取得方法の選択
# "file_name": ファイル名からモードを取得（従来の方法）
# "file_inside": ファイル内（シート内）からモードを取得
# SELECT_MODE_PICKER = "file_name"　# 一つのファイルに複数のモード記載がシート別である場合の対応ができない。
SELECT_MODE_PICKER = "file_inside"

# 抽出する項目名の定義（Excelシート内での名前）
FIELD_NAMES = [
    "車名",
    "通称名",
    "型式",
    "燃費値\n（km/L）",  # 改行が入っている場合もあるため、両方対応
]

# 燃費値の別名（改行なし）
FUEL_EFFICIENCY_ALIASES = [
    "燃費値（km/L）",
    "燃費値\n（km/L）",
    "燃費値\n(km/L)",
    "燃費値(km/L)",
    "(km/L)",
    "(km/L）",
]

# CSV出力時のヘッダー名（日本語のまま）
OUTPUT_HEADERS = [
    "フォルダ名",
    "ファイル名",
    "車名",
    "通称名",
    "型式",
    "燃費値（km/L）",
    "燃料種別",
]

# ファイル名フィルタとCSV出力ファイル名のマッピング
FILE_FILTERS = {
    "WLTC": "output_WLTC.csv",
    "JC08": "output_JC08.csv",
    "10・15": "output_10-15.csv",
    "JH25": "output_JH25.csv",
    "JH15": "output_JH15.csv",
}

# モード名の正規化マッピング（全角英数字版も含む）
# ※normalize_mode_text()で全角→半角変換されるため、
#   基本的には半角版のみで十分だが、念のため全角版も記載
MODE_MAPPING = {
    "WLTCモード": "WLTC",
    "WLTC": "WLTC",
    "ＷＬＴＣモード": "WLTC",  # 全角アルファベット版
    "ＷＬＴＣ": "WLTC",  # 全角アルファベット版
    "JC08モード": "JC08",
    "JC08": "JC08",
    "JC０８モード": "JC08",  # 全角数字版
    "JC０８": "JC08",  # 全角数字版
    "ＪＣ08モード": "JC08",  # 全角アルファベット版
    "ＪＣ08": "JC08",  # 全角アルファベット版
    "ＪＣ０８モード": "JC08",  # 全角英数字版
    "ＪＣ０８": "JC08",  # 全角英数字版
    "10･15モード": "10・15",
    "10・15モード": "10・15",
    "10-15モード": "10・15",
    "10・15": "10・15",
    "10･15": "10・15",
    "１０・１５モード": "10・15",  # 全角数字版
    "１０・１５": "10・15",  # 全角数字版
    "１０･１５モード": "10・15",  # 全角数字版
    "１０･１５": "10・15",  # 全角数字版
    "JH15モード": "JH15",
    "JH15": "JH15",
    "JH１５モード": "JH15",  # 全角数字版
    "JH１５": "JH15",  # 全角数字版
    "ＪＨ15モード": "JH15",  # 全角アルファベット版
    "ＪＨ15": "JH15",  # 全角アルファベット版
    "ＪＨ１５モード": "JH15",  # 全角英数字版
    "ＪＨ１５": "JH15",  # 全角英数字版
    "重車両モード": "JH15",
    "JH25モード": "JH25",
    "JH25": "JH25",
    "JH２５モード": "JH25",  # 全角数字版
    "JH２５": "JH25",  # 全角数字版
    "ＪＨ25モード": "JH25",  # 全角アルファベット版
    "ＪＨ25": "JH25",  # 全角アルファベット版
    "ＪＨ２５モード": "JH25",  # 全角英数字版
    "ＪＨ２５": "JH25",  # 全角英数字版
}

# データ抽出対象の行範囲（4～10行目まで広げて対応）
TARGET_ROWS = range(4, 11)  # Excelは1-indexed、Pythonは0-indexed なので注意

# 検索対象のフォルダ
BASE_FOLDER = "."

# =====================================


def normalize_mode_text(text: str) -> str:
    """
    モード名のテキストを正規化（全角英数字・記号を半角に変換）

    Args:
        text: 正規化するテキスト

    Returns:
        正規化されたテキスト
    """
    if not text:
        return text

    result = text

    # 全角アルファベット（大文字）を半角に変換
    zenkaku_upper = "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
    hankaku_upper = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for z, h in zip(zenkaku_upper, hankaku_upper):
        result = result.replace(z, h)

    # 全角アルファベット（小文字）を半角に変換
    zenkaku_lower = "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
    hankaku_lower = "abcdefghijklmnopqrstuvwxyz"
    for z, h in zip(zenkaku_lower, hankaku_lower):
        result = result.replace(z, h)

    # 全角数字を半角に変換
    zenkaku_nums = "０１２３４５６７８９"
    hankaku_nums = "0123456789"
    for z, h in zip(zenkaku_nums, hankaku_nums):
        result = result.replace(z, h)

    # 全角記号を半角に変換
    result = result.replace("・", "・")  # 中点は既に半角の場合もある
    result = result.replace("･", "・")  # 半角カタカナ中点も統一
    result = result.replace("－", "-")  # 全角ハイフン
    result = result.replace("ー", "-")  # 長音記号

    return result


def extract_fuel_type(sheet, file_path: str, reader: "ExcelReader", mode: str) -> str:
    """
    シートから燃料種別を抽出

    Args:
        sheet: Excelシートオブジェクト
        file_path: ファイルパス
        reader: ExcelReaderインスタンス
        mode: 検出されたモード（JH15/JH25の場合のデフォルト値判定用）

    Returns:
        燃料種別（ガソリン、ディーゼル、LPガス）
    """
    # A1～C5までの範囲を探索（A2セルに燃料情報がある場合もある）
    search_cells = [
        (0, 0),
        (0, 1),
        (0, 2),  # A1, B1, C1
        (1, 0),
        (1, 1),
        (1, 2),  # A2, B2, C2 ← ここに燃料情報があることが多い
        (2, 0),
        (2, 1),
        (2, 2),  # A3, B3, C3
        (3, 0),
        (3, 1),
        (3, 2),  # A4, B4, C4
        (4, 0),
        (4, 1),
        (4, 2),  # A5, B5, C5
    ]

    for row, col in search_cells:
        try:
            value = reader.get_cell_value(sheet, row, col, file_path)
            if not value:
                continue

            value_str = str(value).strip()

            # 全角→半角変換（LPガスの全角対応）
            value_normalized = value_str
            value_normalized = value_normalized.replace("ＬＰ", "LP")
            value_normalized = value_normalized.replace("ｌｐ", "lp")

            # 燃料種別を判定
            if "ガソリン" in value_normalized:
                return "ガソリン"
            elif "ディーゼル" in value_normalized or "軽油" in value_normalized:
                return "ディーゼル"
            elif "LP" in value_normalized.upper() or "LPガス" in value_str:
                return "LPガス"

        except Exception:
            continue

    # シート内で見つからなかった場合、ファイル名から推測
    file_name = Path(file_path).name
    if "ガソリン" in file_name:
        return "ガソリン"
    elif "ディーゼル" in file_name or "軽油" in file_name:
        return "ディーゼル"
    elif "LP" in file_name or "LPガス" in file_name:
        return "LPガス"

    # JH15/JH25モードで燃料が見つからない場合はディーゼル
    if mode in ["JH15", "JH25"]:
        return "ディーゼル"

    # それ以外で見つからない場合は空文字
    return ""


class ExcelReader:
    """Excelファイルを読み込むクラス（xls/xlsx両対応）"""

    @staticmethod
    def read_excel(file_path: str):
        """
        Excelファイルを読み込む
        拡張子に応じてxlsxまたはxls形式で読み込む
        """
        if file_path.endswith(".xlsx"):
            # openpyxlの警告をキャッチ
            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                # 条件付き書式の警告が出た場合、説明を追加
                for warning in w:
                    if "Conditional Formatting" in str(warning.message):
                        print("    ℹ️  条件付き書式を検出（データ読取に影響なし）")
                        break
                return workbook
        elif file_path.endswith(".xls"):
            return xlrd.open_workbook(file_path, formatting_info=False)
        else:
            raise ValueError(f"Unsupported file format: {file_path}")

    @staticmethod
    def get_sheets(workbook, file_path: str):
        """ワークブックからシート一覧を取得"""
        if file_path.endswith(".xlsx"):
            return workbook.worksheets
        else:  # xls
            return workbook.sheets()

    @staticmethod
    def get_cell_value(sheet, row: int, col: int, file_path: str):
        """
        セルの値を取得（結合セルにも対応）
        row, col は0-indexed
        """
        if file_path.endswith(".xlsx"):
            # openpyxl は 1-indexed
            cell = sheet.cell(row + 1, col + 1)
            return cell.value if cell.value is not None else ""
        else:  # xls
            try:
                value = sheet.cell_value(row, col)
                return value if value != "" else ""
            except Exception:
                return ""


def extract_mode_from_sheet(
    sheet, file_path: str, reader: ExcelReader, file_name: str = ""
) -> Tuple[str, bool]:
    """
    シート内からモード名を抽出（徹底的に探索）

    Args:
        sheet: Excelシートオブジェクト
        file_path: ファイルパス
        reader: ExcelReaderインスタンス
        file_name: ファイル名（ログ出力用）

    Returns:
        (モード名, デフォルト使用フラグ)のタプル
        デフォルト使用フラグ: Trueの場合はモードが見つからなかった
    """
    # デバッグモード（特定のファイル名を含む場合に詳細ログ）
    debug_mode = "LPG" in file_name and "WLTC" in file_name

    # シートの最大列数と行数を取得
    if file_path.endswith(".xlsx"):
        max_col = sheet.max_column
        max_row = min(sheet.max_row, 15)  # 最大15行目までチェック
    else:
        max_col = sheet.ncols
        max_row = min(sheet.nrows, 15)  # 最大15行目までチェック

    if debug_mode:
        print(f"      [DEBUG] Scanning {max_row} rows x {max_col} cols")

    # 戦略1: 上位10行×全列を徹底的にスキャン
    for row in range(0, min(10, max_row)):  # 1～10行目
        for col in range(max_col):
            value = reader.get_cell_value(sheet, row, col, file_path)
            if value:
                value_str = str(value).strip()
                # 全角数字を半角に正規化
                value_normalized = normalize_mode_text(value_str)

                # デバッグ出力
                if debug_mode and (
                    "モード" in value_str
                    or "WLTC" in value_str
                    or "ＷＬＴＣ" in value_str
                    or "JC08" in value_str
                    or "ＪＣ０８" in value_str
                    or "JC" in value_str
                    or "ＪＣ" in value_str
                ):
                    print(f"      [DEBUG] ({row + 1},{col + 1}): {value_str[:50]}")
                    if value_str != value_normalized:
                        print(f"      [DEBUG] Normalized: {value_normalized[:50]}")

                # モード名が含まれているかチェック（正規化後の文字列で）
                for mode_key, mode_value in MODE_MAPPING.items():
                    if mode_key in value_normalized:
                        print(
                            f"      Found mode ({row + 1},{col + 1}): "
                            f"{value_str[:30]} -> {mode_value}"
                        )
                        return mode_value, False

    # 戦略2: 「燃費値(km/L)」列を探して、その周辺をチェック
    fuel_col = None
    for check_row in range(0, min(10, max_row)):
        for col_idx in range(max_col):
            cell_value = reader.get_cell_value(sheet, check_row, col_idx, file_path)
            if cell_value:
                cell_str = str(cell_value).strip()
                # 「燃費値」と「km/L」の両方が含まれているかチェック
                if "燃費値" in cell_str and "km/L" in cell_str:
                    fuel_col = col_idx
                    # 燃費値列周辺（上下左右）をチェック
                    for offset_row in range(-2, 5):  # 2行上～4行下
                        for offset_col in range(-5, 6):  # 5列左～5列右
                            check_row_idx = check_row + offset_row
                            check_col_idx = col_idx + offset_col
                            if (
                                0 <= check_row_idx < max_row
                                and 0 <= check_col_idx < max_col
                            ):
                                val = reader.get_cell_value(
                                    sheet, check_row_idx, check_col_idx, file_path
                                )
                                if val:
                                    val_str = str(val).strip()
                                    # 全角数字を半角に正規化
                                    val_normalized = normalize_mode_text(val_str)
                                    for m_key, m_val in MODE_MAPPING.items():
                                        if m_key in val_normalized:
                                            print(
                                                f"      Found mode near fuel "
                                                f"({check_row_idx + 1},"
                                                f"{check_col_idx + 1}): "
                                                f"{val_str[:30]} -> {m_val}"
                                            )
                                            return m_val, False
                    break
        if fuel_col is not None:
            break

    # 戦略3: 特定のキーワードパターンで探索（柔軟なマッチング）
    mode_patterns = [
        "WLTC",
        "WLTCモード",
        "JC08",
        "JC08モード",
        "10・15",
        "10･15",
        "10-15",
        "10・15モード",
        "10･15モード",
        "JH25",
        "JH25モード",
        "JH15",
        "JH15モード",
        "重車両モード",
    ]

    if debug_mode:
        print(f"      [DEBUG] Strategy 3: Pattern matching (rows 0-{min(15, max_row)})")

    for row in range(0, min(15, max_row)):
        for col in range(max_col):
            value = reader.get_cell_value(sheet, row, col, file_path)
            if value:
                # スペースを削除して正規化
                value_str = str(value).strip().replace(" ", "").replace("　", "")
                # 全角数字を半角に正規化
                value_normalized = normalize_mode_text(value_str)
                for pattern in mode_patterns:
                    pattern_norm = pattern.replace(" ", "").replace("　", "")
                    if pattern_norm in value_normalized:
                        mode_value = MODE_MAPPING.get(pattern, pattern)
                        if mode_value in [
                            "WLTC",
                            "JC08",
                            "10・15",
                            "JH25",
                            "JH15",
                        ]:
                            print(
                                f"      Found mode pattern "
                                f"({row + 1},{col + 1}): "
                                f"{value_str[:30]} -> {mode_value}"
                            )
                            return mode_value, False

    # モードが見つからない場合はJH15モード（重車両モード）をデフォルト
    if debug_mode:
        print("      [DEBUG] No mode found, defaulting to JH15")
    print("      Mode not found in sheet, using default: JH15")
    return "JH15", True  # Trueはデフォルト使用を示す


def extract_data_from_sheet(
    sheet,
    file_path: str,
    folder_name: str,
    file_name: str,
    sheet_name: str = "",
    mode: str = None,
) -> Tuple[List[Dict[str, Any]], str, bool]:
    """
    シートから指定された行範囲のデータを抽出

    Args:
        sheet: Excelシートオブジェクト
        file_path: ファイルパス
        folder_name: フォルダ名
        file_name: ファイル名
        sheet_name: シート名
        mode: モード名（指定されている場合）

    Returns:
        (抽出されたデータのリスト, モード名, デフォルト使用フラグ)のタプル
    """
    reader = ExcelReader()
    results = []
    is_default_mode = False

    # モードが指定されていない場合はシートから抽出
    if mode is None:
        if SELECT_MODE_PICKER == "file_inside":
            mode, is_default_mode = extract_mode_from_sheet(
                sheet, file_path, reader, file_name
            )
        else:
            # file_name モードでもモードが指定されていない場合はデフォルト
            mode = "JH15"
            is_default_mode = True

    # 燃料種別を抽出（シート全体から1回だけ取得）
    fuel_type = extract_fuel_type(sheet, file_path, reader, mode)

    # シート内の最大列数と行数を取得
    if file_path.endswith(".xlsx"):
        max_col = sheet.max_column
        max_row = sheet.max_row
    else:
        max_col = sheet.ncols
        max_row = sheet.nrows

    # ヘッダー行（4～10行目）から列のインデックスを特定
    col_indices = {}
    header_last_row = 6  # デフォルトは6行目（7行目からデータが始まる想定）

    for row_idx in TARGET_ROWS:
        for col_idx in range(max_col):
            cell_value = reader.get_cell_value(sheet, row_idx - 1, col_idx, file_path)

            if cell_value:
                cell_str = str(cell_value).strip()

                # 各項目の列インデックスを記録
                if "車名" == cell_str and "車名" not in col_indices:
                    col_indices["車名"] = col_idx
                    header_last_row = max(header_last_row, row_idx)

                if "通称名" == cell_str and "通称名" not in col_indices:
                    # 通称名はヘッダーの次の列にデータがある可能性があるため、両方を試す
                    col_indices["通称名"] = col_idx
                    col_indices["通称名_alt"] = col_idx + 1  # 代替列
                    header_last_row = max(header_last_row, row_idx)

                if "型式" == cell_str and "型式" not in col_indices:
                    col_indices["型式"] = col_idx
                    header_last_row = max(header_last_row, row_idx)

                # 燃費値（改行の有無に対応、JH15形式にも対応）
                # 「燃費値」という文字列が含まれているか、または「(km/L)」だけの場合
                # ただし、「基準値」や他の文字列を含まない単独の「(km/L)」を優先
                if "燃費値（km/L）" not in col_indices:
                    if "燃費値" in cell_str and "（km/L）" in cell_str:
                        # 「燃費値（km/L）」形式
                        col_indices["燃費値（km/L）"] = col_idx
                        header_last_row = max(header_last_row, row_idx)
                    elif cell_str in ["(km/L)", "(km/L）"] or (
                        cell_str.startswith("(km/L") and len(cell_str) <= 10
                    ):
                        # 単独の「(km/L)」形式（JH15など）
                        col_indices["燃費値（km/L）"] = col_idx
                        header_last_row = max(header_last_row, row_idx)

    # ヘッダーが見つからない場合はデフォルトの列位置を使用
    # （一部のファイルはヘッダーに「車名」「通称名」という文字列がないため）
    if "車名" not in col_indices:
        col_indices["車名"] = 0  # Col1をデフォルトの車名列とする
        header_last_row = 6  # デフォルト

    if "通称名" not in col_indices:
        col_indices["通称名"] = 1  # Col2をデフォルトの通称名列とする
        col_indices["通称名_alt"] = 2  # Col3も試す
        header_last_row = 6  # デフォルト

    if "型式" not in col_indices:
        col_indices["型式"] = 3  # Col4をデフォルトの型式列とする
        header_last_row = 6  # デフォルト

    if "燃費値（km/L）" not in col_indices:
        # 燃費値が見つからない場合、後続の処理で探す
        # とりあえず一般的な位置を試す
        for try_col in [9, 10, 11, 12, 13]:  # Col10～14あたりを試す
            if try_col < max_col:
                # ヘッダー行で「km/L」を含む列を探す
                for check_row in range(3, 8):
                    check_val = reader.get_cell_value(
                        sheet, check_row, try_col, file_path
                    )
                    if check_val and (
                        "km/L" in str(check_val) or "km／L" in str(check_val)
                    ):
                        col_indices["燃費値（km/L）"] = try_col
                        break
                if "燃費値（km/L）" in col_indices:
                    break

    # データ行（ヘッダーの次の行以降）からデータを抽出
    # ヘッダーの最後の行の次からデータが始まる
    # ただし、データが実際にどこから始まるかを確認する
    data_start_row = (
        header_last_row + 2
    )  # ヘッダーの次の行（+1）、さらに1行スペース（+1）の可能性を考慮

    # データ開始行を探す（型式列に値がある最初の行）
    if "型式" in col_indices:
        for search_row in range(7, min(15, max_row)):  # 8～15行目あたりを探す
            type_value = reader.get_cell_value(
                sheet, search_row, col_indices["型式"], file_path
            )
            if (
                type_value
                and str(type_value).strip()
                and str(type_value).strip() not in ["型式", "None"]
            ):
                data_start_row = search_row + 1  # 0-indexed なので+1
                break

    # 結合セル対応：直前の値を保持する変数
    last_values = {}

    for row_idx in range(data_start_row - 1, max_row):
        row_data = {}
        row_data["フォルダ名"] = folder_name
        row_data["ファイル名"] = file_name
        row_data["燃料種別"] = fuel_type  # 燃料種別を追加

        # 各項目の値を取得
        for field_name, col_idx in col_indices.items():
            # 代替列は後で処理するのでスキップ
            if field_name.endswith("_alt"):
                continue

            value = reader.get_cell_value(sheet, row_idx, col_idx, file_path)

            # 通称名の特別処理
            if field_name == "通称名":
                value_str = str(value).strip() if value else ""

                # 改行をスペースに置換してから判定（記号チェック用）
                # これにより『※\n※』が『※ ※』になる
                value_normalized = value_str.replace("\n", " ").replace("\r", " ")
                value_normalized = " ".join(
                    value_normalized.split()
                )  # 連続スペースを1つに

                # 通称名が空、または記号のみの場合は代替列をチェック
                # 半角記号と全角記号の両方に対応
                symbol_patterns = [
                    "※1",
                    "※",
                    "＊",
                    "*2",
                    "*",
                    "※2",
                    "＊1",
                    "※１",
                    "※２",
                    "＊１",
                    "＊２",  # 全角数字も追加
                    "※3",
                    "※３",
                    "＊３",
                    "*3",
                ]

                # スペース区切りの記号もチェック（例：『※ ※』『※1 ※2』など）
                is_symbol_only = False
                if value_normalized and " " in value_normalized:
                    parts = value_normalized.split()
                    if all(part in symbol_patterns for part in parts):
                        is_symbol_only = True
                elif value_normalized in symbol_patterns:
                    is_symbol_only = True

                if not value_normalized or value_normalized == "None" or is_symbol_only:
                    # 代替列をチェック
                    if "通称名_alt" in col_indices:
                        alt_value = reader.get_cell_value(
                            sheet, row_idx, col_indices["通称名_alt"], file_path
                        )
                        alt_str = str(alt_value).strip() if alt_value else ""
                        if (
                            alt_str
                            and alt_str != "None"
                            and alt_str not in symbol_patterns
                        ):
                            value = alt_value

            # 値の処理
            value_str = str(value).strip() if value else ""

            if value_str and value_str != "None":
                # 値がある場合は使用し、記憶する
                row_data[field_name] = value_str
                # 通称名の場合は、型式と同じでないか確認してから記憶する
                if field_name == "通称名":
                    # 型式フォーマット（XXX-XXXXXX）の場合は記憶しない
                    is_katashiki_format = (
                        len(value_str) >= 4
                        and value_str[3:4] == "-"
                        and "-" in value_str
                    )
                    if not is_katashiki_format:
                        last_values[field_name] = value_str
                # 車名の場合のみ記憶
                elif field_name == "車名":
                    last_values[field_name] = value_str
            elif field_name == "通称名":
                # 通称名が空の場合は前の値を使用（結合セル・連続レコード対応）
                if "通称名" in last_values and last_values["通称名"]:
                    row_data[field_name] = last_values[field_name]
                else:
                    row_data[field_name] = ""
            elif field_name == "車名":
                # 車名が空の場合
                if "車名" in last_values and last_values["車名"]:
                    # 前の値を使用
                    row_data[field_name] = last_values[field_name]
                elif sheet_name:
                    # シート名を使用
                    row_data[field_name] = sheet_name
                    last_values[field_name] = sheet_name
                else:
                    row_data[field_name] = ""
            elif value_str:
                # その他の項目
                row_data[field_name] = value_str
            else:
                # 値が空の場合、空文字列を設定
                row_data[field_name] = ""

        # 型式があればデータとして扱う（型式は結合されない主キー的な項目）
        if "型式" in row_data and row_data["型式"]:
            # 注釈行をスキップ
            is_note_row = False

            # 1. 型式フォーマットチェック（4文字目に-がないものは注釈行）
            katashiki = str(row_data["型式"]).strip()
            if len(katashiki) >= 4 and katashiki[3:4] != "-":
                # 型式の正しいフォーマット: XXX-XXXXXX
                # 4文字目が-でない場合は注釈行の可能性が高い
                is_note_row = True

            # 2. 通称名や型式に注釈パターンが含まれているかチェック
            if not is_note_row:
                check_fields = ["通称名", "型式", "燃費値（km/L）"]
                note_patterns = [
                    "（注）",
                    "(注)",
                    "※印",
                    "*印",
                    "注：",
                    "製造事業者",
                    "Bayerische",
                    "Motoren",
                    "付いている通称名",
                    "については",
                    "実際に販売されている",
                    "異なる場合があります",
                    "記載された車両重量",
                    "車両総重量",
                ]

                for field in check_fields:
                    if field in row_data:
                        value = str(row_data[field])
                        for pattern in note_patterns:
                            if pattern in value:
                                is_note_row = True
                                break
                        if is_note_row:
                            break

            # 3. JH15の通称名問題を修正：通称名が型式と同じ場合は結合セルとして扱う
            if not is_note_row and "通称名" in row_data:
                tsusho = str(row_data.get("通称名", "")).strip()
                # 通称名が型式と同じ、または型式フォーマットの場合
                if tsusho and (
                    tsusho == katashiki
                    or (len(tsusho) >= 4 and tsusho[3:4] == "-" and "-" in tsusho)
                ):
                    # 結合セル対応：前の値を使用
                    if "通称名" in last_values and last_values["通称名"] != tsusho:
                        row_data["通称名"] = last_values["通称名"]
                    else:
                        # 前の値もない場合は空にする（シート名などで補完される）
                        row_data["通称名"] = ""

            # 注釈行でなければ追加
            if not is_note_row:
                results.append(row_data)
        # 連続して空行が続いたらデータの終わりと判断
        elif row_idx > data_start_row + 10:  # データ開始後10行空行があったら終了
            # 全ての列が空かチェック
            all_empty = True
            for col_idx in range(max_col):
                value = reader.get_cell_value(sheet, row_idx, col_idx, file_path)
                if value and str(value).strip() and str(value).strip() != "None":
                    all_empty = False
                    break
            if all_empty:
                # さらに次の5行も空かチェック
                empty_count = 0
                for check_row in range(row_idx, min(row_idx + 5, max_row)):
                    row_empty = True
                    for col_idx in range(max_col):
                        value = reader.get_cell_value(
                            sheet, check_row, col_idx, file_path
                        )
                        if (
                            value
                            and str(value).strip()
                            and str(value).strip() != "None"
                        ):
                            row_empty = False
                            break
                    if row_empty:
                        empty_count += 1

                if empty_count >= 3:  # 3行以上連続して空行ならデータ終了
                    # 結合セル用の記憶をクリア
                    last_values = {}
                    break

    return results, mode, is_default_mode


def get_primary_mode_from_filename(file_name: str) -> str:
    """
    ファイル名から最も優先度の高いモードを取得

    Args:
        file_name: ファイル名

    Returns:
        モード名（優先度順：WLTC > JC08 > 10・15 > JH25 > JH15）
    """
    # 優先順位が高い順にチェック
    mode_priority = ["WLTC", "JC08", "10・15", "10･15", "JH25", "JH15"]

    for mode in mode_priority:
        if mode in file_name:
            # 10・15と10･15は同じモードとして扱う
            if mode in ["10・15", "10･15"]:
                return "10・15"
            return mode

    return None


def process_excel_files(
    base_folder: str, filter_keyword: str = None
) -> Tuple[Dict[str, List[Dict[str, Any]]], List[Tuple[str, str, str]]]:
    """
    指定フォルダ内のExcelファイルを処理

    Args:
        base_folder: 基準フォルダ
        filter_keyword: ファイル名フィルタキーワード（file_name モードの場合のみ使用）

    Returns:
        (モード別のデータ辞書, モード未検出ファイルリスト)のタプル
        モード未検出ファイルリスト: [(フォルダ名, ファイル名, シート名), ...]
    """
    all_data = {}
    base_path = Path(base_folder)
    no_mode_files = []  # モードが見つからなかったファイルのリスト

    # モード別にデータを初期化
    for mode in ["WLTC", "JC08", "10・15", "JH25", "JH15"]:
        all_data[mode] = []

    # フォルダを走査
    for subfolder in base_path.iterdir():
        if not subfolder.is_dir():
            continue

        folder_name = subfolder.name
        print(f"Processing folder: {folder_name}")

        # フォルダ内のExcelファイルを走査
        for file_path in subfolder.glob("*"):
            if file_path.suffix not in [".xls", ".xlsx"]:
                continue

            file_name = file_path.name

            # file_name モードの場合はファイル名フィルタを適用
            if SELECT_MODE_PICKER == "file_name" and filter_keyword:
                # ファイル名から最も優先度の高いモードを取得
                primary_mode = get_primary_mode_from_filename(file_name)

                # 現在のフィルタと一致する場合のみ処理
                if primary_mode != filter_keyword:
                    continue

                current_mode = filter_keyword
            else:
                current_mode = None  # file_inside モードの場合はシートから取得

            print(f"  Processing file: {file_name}")

            try:
                # Excelファイルを開く
                workbook = ExcelReader.read_excel(str(file_path))
                sheets = ExcelReader.get_sheets(workbook, str(file_path))

                # 各シートを処理
                for sheet in sheets:
                    try:
                        # シート名を取得
                        if str(file_path).endswith(".xlsx"):
                            sheet_name = sheet.title if hasattr(sheet, "title") else ""
                        else:  # xls
                            sheet_name = sheet.name if hasattr(sheet, "name") else ""

                        print(f"    Processing sheet: {sheet_name}")
                        data, mode, is_default = extract_data_from_sheet(
                            sheet,
                            str(file_path),
                            folder_name,
                            file_name,
                            sheet_name,
                            current_mode,
                        )

                        # シート処理結果をログ出力
                        print(
                            f"      -> Detected mode: {mode}, Data count: {len(data)}"
                        )

                        # デフォルトモードを使用した場合は記録
                        if is_default:
                            no_mode_files.append((folder_name, file_name, sheet_name))

                        # モード別にデータを格納
                        if mode in all_data:
                            all_data[mode].extend(data)
                            print(f"      -> Added to {mode} data")
                        else:
                            print(
                                f"      Warning: Unknown mode '{mode}', skipping data"
                            )
                    except Exception as e:
                        print(f"    Error processing sheet in {file_name}: {e}")
                        import traceback

                        traceback.print_exc()

                # ワークブックを閉じる
                if str(file_path).endswith(".xls"):
                    workbook.release_resources()

            except Exception as e:
                print(f"    Error processing file {file_name}: {e}")
                import traceback

                traceback.print_exc()

    return all_data, no_mode_files


def normalize_katashiki(text: str) -> str:
    """
    型式の文字列を正規化（全角→半角変換、改行削除）

    Args:
        text: 型式の文字列

    Returns:
        正規化された文字列
    """
    if not text:
        return text

    # 改行を削除（型式は1行にまとめる）
    result = text.replace("\n", "").replace("\r", "")

    # 全角英数字を半角に変換
    zenkaku = "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ０１２３４５６７８９"
    hankaku = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"

    for z, h in zip(zenkaku, hankaku):
        result = result.replace(z, h)

    # 全角ハイフンを半角に
    result = result.replace("－", "-").replace("ー", "-")

    # 先頭・末尾の空白を削除
    result = result.strip()

    return result


def normalize_tsusho(text: str) -> str:
    """
    通称名の文字列を正規化（改行削除、連続スペース削除、全角括弧を半角に）

    Args:
        text: 通称名の文字列

    Returns:
        正規化された文字列
    """
    if not text:
        return text

    import re

    # 改行を空白に置換
    result = text.replace("\n", " ").replace("\r", " ")

    # 全角括弧を半角に変換
    result = result.replace("（", "(").replace("）", ")")

    # 連続する空白を1つにまとめる
    result = re.sub(r"\s+", " ", result)

    # 先頭・末尾の空白を削除
    result = result.strip()

    return result


def write_to_csv(data: List[Dict[str, Any]], output_file: str):
    """
    データをCSVファイルに書き込む

    Args:
        data: 出力データのリスト
        output_file: 出力ファイル名
    """
    if not data:
        print(f"No data to write to {output_file}")
        return

    with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()

        for row in data:
            # ヘッダーにない項目は無視し、ヘッダーにある項目で値がないものは空文字を設定
            output_row = {}
            for header in OUTPUT_HEADERS:
                value = row.get(header, "")
                # 型式は全角→半角変換、改行削除
                if header == "型式" and value:
                    value = normalize_katashiki(value)
                # 通称名は改行削除、連続スペース削除
                elif header == "通称名" and value:
                    value = normalize_tsusho(value)
                output_row[header] = value
            writer.writerow(output_row)

    print(f"Wrote {len(data)} rows to {output_file}")


def main():
    """メイン処理"""
    print("Starting Excel consolidation...")
    print(f"Base folder: {BASE_FOLDER}")
    print(f"Mode picker: {SELECT_MODE_PICKER}")
    print()

    all_no_mode_files = []  # モード未検出ファイルの集約リスト

    if SELECT_MODE_PICKER == "file_name":
        # ファイル名からモードを取得する従来の方法
        print("Mode detection: from filename")
        print(f"Target filters: {list(FILE_FILTERS.keys())}")
        print()

        # 各フィルタに対して処理を実行
        for filter_keyword, output_file in FILE_FILTERS.items():
            print(f"\n{'=' * 60}")
            print(f"Processing files with '{filter_keyword}' in filename")
            print(f"{'=' * 60}")

            mode_data, no_mode_files = process_excel_files(BASE_FOLDER, filter_keyword)
            all_no_mode_files.extend(no_mode_files)
            if filter_keyword in mode_data:
                write_to_csv(mode_data[filter_keyword], output_file)
            else:
                print(f"No data found for mode {filter_keyword}")

    elif SELECT_MODE_PICKER == "file_inside":
        # ファイル内からモードを取得する新しい方法
        print("Mode detection: from file inside")
        print()

        print(f"\n{'=' * 60}")
        print("Processing all Excel files...")
        print(f"{'=' * 60}")

        # 全てのファイルを処理（フィルタなし）
        mode_data, no_mode_files = process_excel_files(BASE_FOLDER)
        all_no_mode_files = no_mode_files

        # モード別にCSV出力
        for mode, output_file in FILE_FILTERS.items():
            if mode in mode_data and mode_data[mode]:
                write_to_csv(mode_data[mode], output_file)
            else:
                print(f"No data found for mode {mode}")

    else:
        print(f"Error: Unknown SELECT_MODE_PICKER value: {SELECT_MODE_PICKER}")
        return

    # モード未検出ファイルの表示
    if all_no_mode_files:
        print("\n" + "=" * 60)
        print("⚠️  モードが見つからなかったファイル（デフォルトJH15使用）")
        print("=" * 60)
        for folder, file, sheet in all_no_mode_files:
            print(f"  フォルダ: {folder}")
            print(f"  ファイル: {file}")
            if sheet:
                print(f"  シート: {sheet}")
            print()

    print("\n" + "=" * 60)
    print("All processing completed!")
    print(f"Total files with mode not found: {len(all_no_mode_files)}")
    print("=" * 60)


if __name__ == "__main__":
    main()
